import openpyxl
import traceback
import os
import json
import signal
import sys

RED = "\033[91m"
YELLOW = "\033[93m"
GREEN = "\033[92m"
WHITE = "\033[97m"
RESET = "\033[0m"

def load_classes(sheet):
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {
            'Course Code': row[0],
            'Course Name': row[1],
            'Section': row[2],
            'Full Course Name': row[3],
            'Instructor': row[4],
            'Schedule': row[5],
            'Location': row[6],
            'Capacity': row[7]
        }
        data.append(row_data)
    return data

def find_excel_file():
    excel_files = [f for f in os.listdir('.') if os.path.isfile(f) and (f.endswith('.xlsx') or f.endswith('.xls'))]
    
    if excel_files:
        print("Excel files found in the current directory:")
        for idx, file in enumerate(excel_files, start=1):
            print(f"{idx}. {file}")
        print(f"{len(excel_files) + 1}. Select path manually")
        
        try:
            choice = int(input("Choose an option (enter the number): ").strip())
            if 1 <= choice <= len(excel_files):
                file_name = excel_files[choice - 1]
            elif choice == len(excel_files) + 1:
                file_name = input("Enter the path to the Excel file: ").strip()
            else:
                print("Invalid choice.")
                exit()
        except ValueError:
            print("Invalid input. Please enter a number.")
            exit()
    else:
        file_name = input("No Excel files found in the current directory. Enter the path to the Excel file: ").strip()
    
    if not os.path.exists(file_name):
        print("File not found.")
        exit()
    
    return file_name

def list_sheets(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name, data_only=True)
        return workbook.sheetnames
    except Exception as e:
        print(f"Error loading file: {e}")
        return []

def load_sheet(file_name, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_name, data_only=True)
        return workbook[sheet_name]
    except Exception as e:
        print(f"Error loading sheet: {e}")
        return None

def get_sheet(file_name):
    sheets = list_sheets(file_name)
    if not sheets:
        print("No sheets found or unable to load workbook.")
        return
    
    print("Available sheets:")
    for i, sheet_name in enumerate(sheets, start=1):
        print(f"{i}. {sheet_name}")
    
    while True:
        try:
            sheet_index = int(input("Select a sheet by number: ")) - 1
            if sheet_index < 0 or sheet_index >= len(sheets):
                print("Invalid selection.")
                continue
            break
        except ValueError:
            print("Invalid input.")
            continue
    
    selected_sheet_name = sheets[sheet_index]
    sheet = load_sheet(file_name, selected_sheet_name)
    if not sheet:
        print("Unable to load the selected sheet.")
        return
    return sheet

def main():
    file_name = find_excel_file()
    
    sheet = get_sheet(file_name)
        
    classList = load_classes(sheet)
    
    print(json.dumps(classList, indent=4))

if __name__ == "__main__":
    main()
